#!/usr/bin/env python
# -*- coding: utf-8 -*-
# From: https://stackoverflow.com/questions/63116307/python-get-input-from-hid-device-outside-the-active-window
# This script reads the output from HID device e.g. ECCO bar/rfid scanner
# Make sure evdev is installed or install it with: pip3 install evdev.
# The Kernel on the Pi creates an input device which resides in /dev/input/.
# You can find out the names and other attributes of different devices using:
# cat /proc/bus/input/devices

"""Gather barcode data from badges to track attendance
   QR code format is specific to Trail Life USA badges"""
import argparse
import logging
import time
from datetime import datetime
from openpyxl import Workbook

import evdev

class AttTrack:
    def __init__(self, logging, vendor, product):
        self.logging = logging
        self.mVendorStr = vendor
        self.mProductStr = product
        self.mPollIntervalSecs = 1
        self.mExclusiveAccess = 1
        self.mDocumentCreated = False
        self.mDeviceGrabbed = False
        self.mBarcodeReader = None
        self.CODE_MAP_CHAR = {
            'KEY_MINUS': "-",
            'KEY_SPACE': " ",    
            'KEY_U': "U",
            'KEY_W': "W",
            'KEY_BACKSLASH': "\\",
            'KEY_GRAVE': "`",
            'KEY_NUMERIC_STAR': "*",
            'KEY_NUMERIC_3': "3",
            'KEY_NUMERIC_2': "2",
            'KEY_NUMERIC_5': "5",
            'KEY_NUMERIC_4': "4",
            'KEY_NUMERIC_7': "7",
            'KEY_NUMERIC_6': "6",
            'KEY_NUMERIC_9': "9",
            'KEY_NUMERIC_8': "8",
            'KEY_NUMERIC_1': "1",
            'KEY_NUMERIC_0': "0",
            'KEY_E': "E",
            'KEY_D': "D",
            'KEY_G': "G",
            'KEY_F': "F",
            'KEY_A': "A",
            'KEY_C': "C",
            'KEY_B': "B",
            'KEY_M': "M",
            'KEY_L': "L",
            'KEY_O': "O",
            'KEY_N': "N",
            'KEY_I': "I",
            'KEY_H': "H",
            'KEY_K': "K",
            'KEY_J': "J",
            'KEY_Q': "Q",
            'KEY_P': "P",
            'KEY_S': "S",
            'KEY_X': "X",
            'KEY_Z': "Z",
            'KEY_q': "q",
            'KEY_w': "w",
            'KEY_e': "e",
            'KEY_r': "r",
            'KEY_t': "t",
            'KEY_z': "z",
            'KEY_u': "u",
            'KEY_i': "i",
            'KEY_o': "o",
            'KEY_p': "p",
            'KEY_a': "a",
            'KEY_s': "s",
            'KEY_d': "d",
            'KEY_f': "f",
            'KEY_g': "g",
            'KEY_h': "h",
            'KEY_j': "k",
            'KEY_l': "l",
            'KEY_y': "y",
            'KEY_x': "x",
            'KEY_c': "c",
            'KEY_v': "v",
            'KEY_b': "b",
            'KEY_n': "n",
            'KEY_m': "m",
            'KEY_KP4': "4",
            'KEY_KP5': "5",
            'KEY_KP6': "6",
            'KEY_KP7': "7",
            'KEY_KP0': "0",
            'KEY_KP1': "1",
            'KEY_KP2': "2",
            'KEY_KP3': "3",
            'KEY_KP8': "8",
            'KEY_KP9': "9",
            'KEY_5': "5",
            'KEY_4': "4",
            'KEY_7': "7",
            'KEY_6': "6",
            'KEY_1': "1",
            'KEY_0': "0",
            'KEY_3': "3",
            'KEY_2': "2",
            'KEY_9': "9",
            'KEY_8': "8",
            'KEY_LEFTBRACE': "[",
            'KEY_RIGHTBRACE': "]",    
            'KEY_COMMA': ",",
            'KEY_EQUAL': "=",    
            'KEY_SEMICOLON': ";",
            'KEY_APOSTROPHE': "'",
            'KEY_T': "T",
            'KEY_V': "V",
            'KEY_R': "R",
            'KEY_Y': "Y",
            'KEY_TAB': "\t",
            'KEY_DOT': ".",
            'KEY_SLASH': "/",
        }
    
    def parse_key_to_char(self, val):
        """Convert key code to character"""
        return self.CODE_MAP_CHAR[val] if val in self.CODE_MAP_CHAR else ""

    def list_devices(self):
        """Log all barcode scanner devices"""
        self.logging.info("List of devices")
        self.logging.info("===============")
        devices = [evdev.InputDevice(path) for path in evdev.list_devices()]
        for aDevice in devices:
            self.logging.info("  {}\t{}\t{}\t{}".format(aDevice.path,
                aDevice.name,
                hex(aDevice.info.vendor),
                hex(aDevice.info.product)))
        self.logging.info("===============")

    def find_device(self):
        """Find USB barcode reader device and grab it"""
        # Wait for device available
        deviceNotFound = True
        self.logging.info("Looking for device")
        while (deviceNotFound):
            # Read list of devices
            devices = [evdev.InputDevice(path) for path in evdev.list_devices()]
            for aDevice in devices:
                if (hex(aDevice.info.vendor) == self.mVendorStr and
                    hex(aDevice.info.product) == self.mProductStr):
                    logMsg = "Found device: " + aDevice.name + " on " + aDevice.path
                    self.logging.info(logMsg)
                    self.mBarcodeReader = evdev.device.InputDevice(aDevice.path)
                    if int(self.mExclusiveAccess) == 1:
                        self.mBarcodeReader.grab()
                        self.mDeviceGrabbed = True
                    return self.mBarcodeReader
                else:
                    time.sleep(self.mPollIntervalSecs)

    def create_document(self, filename):
        """Create spreadsheet for saving attendance data"""
        if not self.mDocumentCreated:
            self.filename = filename
            self.wb = Workbook()
            self.ws = self.wb.create_sheet('Attendance', 0)
            del self.wb['Sheet']
            self.row = 1
            self.logging.info(f'Creating document: {self.filename}')
            self.mDocumentCreated = True

    def add_barcode_to_document(self, barcodeStr):
        """Add barcode data to spreadsheet"""
        # Split barcode into member ID and user ID fields and
        # put in separate columms in the spreadsheet
        idCellRef = f"A{self.row}"
        memberIdCellRef = f"B{self.row}"
        (memberIdStr, idStr) = barcodeStr.split("\\")
        idStr = idStr.strip().lower()
        memberIdStr = memberIdStr.strip()
        self.ws[idCellRef] = idStr
        self.ws[memberIdCellRef] = memberIdStr
        self.row = self.row + 1

    def close_document(self):
        """Close spreadsheet / write to file"""
        if self.mDocumentCreated:
            self.logging.info("Closing file: " + self.filename)
            self.wb.save(self.filename)
            self.mDocumentCreated = False

    def close_device(self, deviceMissing = False):
        """Close USB barcode scanner device"""
        if deviceMissing:
            self.mDeviceGrabbed = False
            if self.mBarcodeReader is not None:
                self.mBarcodeReader.close()
        else: # Device exists, need proper shutdown
            if self.mBarcodeReader is not None:
                if self.mDeviceGrabbed:
                    self.logging.info("Releasing device")
                    self.mBarcodeReader.ungrab()
                    self.mDeviceGrabbed = False
                self.logging.info("Closing device")
                self.mBarcodeReader.close()
        self.mBarcodeReader = None

def main():
    # Process command line input
    parser = argparse.ArgumentParser(
        prog="attTrack",
        description="Gather info from badge QR codes for attendance tracking"
    )
    dt = datetime.now()
    defaultFilename = dt.strftime('%Y%m%d_%H%M%S') + '.xlsx'
    parser.add_argument('-f', '--filename', help='Attendance file', 
        default=defaultFilename)
    parser.add_argument('-l', '--log', help='Log file',
        default='./AttTrack.log')
    parser.add_argument('--vendor', help="USB Vendor String (hex)",
        default="0x26f1")
    parser.add_argument('--product', help="USB Product String (hex)",
        default="0x5651")
    args = parser.parse_args()
    # Set up logging
    logging.basicConfig(level=logging.INFO, filename=args.log)

    # Set up attendance tracking
    myAT = AttTrack(logging, args.vendor, args.product)

    # Log devices at startup
    myAT.list_devices()

    # Main loop for processing barcodes
    continueProcessing = True
    while (continueProcessing):
        reader = myAT.find_device()
        myAT.create_document(args.filename)

        logging.info("Waiting for scan")
        continueScanning = True
        barcodeStr = ""
        while (continueScanning):
            try:
                for event in reader.read_loop():
                    if event.type == evdev.ecodes.EV_KEY:
                        e = evdev.util.categorize(event)
                        if e.keystate == e.key_up:
                            cifra = myAT.parse_key_to_char(e.keycode)
                            barcodeStr = barcodeStr + cifra
                        if e.scancode == 28:
                            if len(barcodeStr) > 0:
                                if (barcodeStr == "948NINJADOWN"):
                                    continueScanning = False
                                    continueProcessing = False
                                    logging.info("Exit code scanned")
                                    myAT.close_document()
                                    myAT.close_device()
                                    break
                                logMsg = "Barcode: [" + barcodeStr + "]"
                                logging.info(logMsg)
                                # Save current barcode
                                myAT.add_barcode_to_document(barcodeStr)
                                barcodeStr = ""
            except FileNotFoundError as ex:
                # Device not found, return to device search
                logging.info("Device not found, returning to search!")
                continueScanning = False
                myAT.close_document()
            except OSError as ex:
                # Device disconnected
                logging.info("Device disconnected")
                continueScanning = False
                myAT.close_device(True)
            except KeyboardInterrupt:
                continueScanning = False
                continueProcessing = False
                logging.info("")
                myAT.close_document()
                myAT.close_device()

if __name__ == "__main__":
    main()

    

